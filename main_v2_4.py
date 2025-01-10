import cx_Oracle
import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime

def construir_query(anio):
    """
    Construye la consulta SQL para el año especificado.
    """
    query_base = """
    SELECT --FORM 01
        t.o_beneficiario AS Beneficiario,
        CAST(t.aa_resolucion AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_resolucion AS VARCHAR2(10)) AS t_formulario,
        CAST(t.o_resolucion AS NUMBER) AS o_formulario,
        TRUNC(t.f_ingreso) AS fh_imputacion,
        TRUNC(d.fh_imputacion) AS NUEVA_FH_IMPUTACION,
        CAST(NULL AS VARCHAR2(10)) AS aa_comprobante,
        CAST(NULL AS VARCHAR2(10)) AS t_comprobante,
        CAST(NULL AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        SUM(d.i_devengado) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tresolucion' AS comentario
    FROM gs_tresolucion t
    JOIN gs_dresol_item d
       ON t.aa_resolucion = d.aa_resolucion
      AND t.t_resolucion = d.t_resolucion
      AND t.o_resolucion = d.o_resolucion
    WHERE t.t_resolucion = 'RSD'
      AND t.e_resolucion = 'Z'
      AND d.c_numcred IS NOT NULL
      AND (t.f_ingreso IS NULL OR t.aa_resolucion != EXTRACT(YEAR FROM t.f_ingreso))
      AND t.aa_resolucion = {anio}
    GROUP BY
        t.o_beneficiario,
        CAST(t.aa_resolucion AS VARCHAR2(10)),
        CAST(t.t_resolucion AS VARCHAR2(10)),
        CAST(t.o_resolucion AS NUMBER),
        TRUNC(t.f_ingreso),
        TRUNC(d.fh_imputacion)
    HAVING NVL(SUM(d.i_devengado), 0) != 0

UNION ALL

SELECT --FORM 02
        t.o_ente AS Beneficiario,
        CAST(t.aa_devengado AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_devengado AS VARCHAR2(10)) AS t_formulario,
        CAST(t.n_devengado AS NUMBER) AS o_formulario,
        TRUNC(di.fh_imputacion) AS fh_imputacion,
        TRUNC(di.fh_imputacion) AS NUEVA_FH_IMPUTACION,
        CAST(NULL AS VARCHAR2(10)) AS aa_comprobante,
        CAST(NULL AS VARCHAR2(10)) AS t_comprobante,
        CAST(NULL AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        SUM(di.i_devengado) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tdevengado' AS comentario
    FROM gs_tdevengado t
    JOIN gs_ddevengado_ffi di
       ON t.o_devengado = di.o_devengado
    WHERE t.e_devengado = 'A'
      AND di.c_numcred IS NOT NULL
      AND t.aa_devengado = {anio}
      AND t.aa_devengado != EXTRACT(YEAR FROM di.fh_imputacion)
    GROUP BY
        t.o_ente,
        CAST(t.aa_devengado AS VARCHAR2(10)),
        CAST(t.t_devengado AS VARCHAR2(10)),
        CAST(t.n_devengado AS NUMBER),
        TRUNC(di.fh_imputacion),
        TRUNC(di.fh_imputacion)
    HAVING NVL(SUM(di.i_devengado), 0) != 0

UNION ALL

SELECT -- FORM 03
        t.o_ente AS Beneficiario,
        CAST(t.aa_precepcion AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_precepcion AS VARCHAR2(10)) AS t_formulario,
        CAST(t.n_precepcion AS NUMBER) AS o_formulario,
        TRUNC(t.fh_imputacion) AS fh_imputacion,
        TRUNC(t.fh_imputacion) AS NUEVA_FH_IMPUTACION,
        CAST(t.aa_ocompra AS VARCHAR2(10)) AS aa_comprobante,
        CAST(t.t_ocompra AS VARCHAR2(10)) AS t_comprobante,
        CAST(t.n_ocompra AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        SUM(NVL(d.i_total, 0)) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tparte_recepcion' AS comentario
    FROM prd_tparte_recepcion t
    JOIN prd_dparte_recepcion_ffi d
       ON t.aa_precepcion = d.aa_precepcion
      AND t.t_precepcion = d.t_precepcion
      AND t.n_precepcion = d.n_precepcion
    WHERE t.e_formulario IN ('A')
      AND t.aa_precepcion = {anio}
      AND (t.fh_imputacion IS NULL OR t.aa_precepcion != EXTRACT(YEAR FROM t.fh_imputacion))
    GROUP BY
        t.o_ente,
        CAST(t.aa_precepcion AS VARCHAR2(10)),
        CAST(t.t_precepcion AS VARCHAR2(10)),
        CAST(t.n_precepcion AS NUMBER),
        TRUNC(t.fh_imputacion),
        TRUNC(t.fh_imputacion),
        CAST(t.aa_ocompra AS VARCHAR2(10)),
        CAST(t.t_ocompra AS VARCHAR2(10)),
        CAST(t.n_ocompra AS NUMBER)
    HAVING NVL(SUM(d.i_total), 0) != 0

UNION ALL

SELECT --FORM 04
        t.o_contratista AS Beneficiario,
        CAST(t.aa_certificado AS VARCHAR2(10)) AS aa_formulario,
        'CAO' AS t_formulario,
        CAST(t.o_certificado AS NUMBER) AS o_formulario,
        TRUNC(t.f_certificacion) AS fh_imputacion,
        TRUNC(d.F_IMPUTACION) AS NEW_FH_IMPUTACION,
        CAST(t.t_contrato AS VARCHAR2(10)) AS aa_comprobante,
        CAST(t.n_contrato AS VARCHAR2(10)) AS t_comprobante,
        CAST(t.aa_contrato AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS NEW_FH_IMPUTACION,
        SUM(d.i_devengado) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tcertif_avance' AS comentario
    FROM obp_tcertificado_avance t
    JOIN obp_dcert_avance_ffi d
       ON t.aa_certificado = d.aa_certificado
      AND t.t_form_medicion = d.t_form_medicion
      AND t.o_certificado = d.o_certificado
      AND t.c_obra = d.co_obra
      AND t.n_dev = d.n_dev
    WHERE t.e_certificado = 'A'
      AND t.aa_certificado = {anio}
      AND (d.f_imputacion IS NULL OR t.aa_certificado != EXTRACT(YEAR FROM d.f_imputacion))
    GROUP BY
        t.o_contratista,
        CAST(t.aa_certificado AS VARCHAR2(10)),
        'CAO',
        CAST(t.o_certificado AS NUMBER),
        TRUNC(t.f_certificacion),
        TRUNC(d.F_IMPUTACION),
        CAST(t.t_contrato AS VARCHAR2(10)),
        CAST(t.n_contrato AS VARCHAR2(10)),
        CAST(t.aa_contrato AS NUMBER)
    HAVING NVL(SUM(d.i_devengado), 0) != 0

UNION ALL

SELECT -- FORM 05
        t.o_ente AS Beneficiario,
        CAST(t.aa_certificado AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_certificado AS VARCHAR2(10)) AS t_formulario,
        CAST(t.n_certificado AS NUMBER) AS o_formulario,
        TRUNC(t.fh_autorizacion) AS fh_imputacion,
        TRUNC(t.FH_AUTORIZACION) AS NEW_FH_IMPUTACION,
        CAST(t.t_ocompra AS VARCHAR2(10)) AS aa_comprobante,
        CAST(t.n_ocompra AS VARCHAR2(10)) AS t_comprobante,
        CAST(t.aa_ocompra AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        SUM(d.i_devengado) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tcertificado' AS comentario
    FROM obp_tcertificado t
    JOIN obp_dcertificado_ffi d
       ON t.aa_certificado = d.aa_certificado
      AND t.t_certificado = d.t_certificado
      AND t.n_certificado = d.n_certificado
    WHERE t.e_certificado = 'A'
      AND t.aa_certificado = {anio}
      AND (t.fh_autorizacion IS NULL OR t.aa_certificado != EXTRACT(YEAR FROM t.fh_autorizacion))
    GROUP BY
        t.o_ente,
        CAST(t.aa_certificado AS VARCHAR2(10)),
        CAST(t.t_certificado AS VARCHAR2(10)),
        CAST(t.n_certificado AS NUMBER),
        TRUNC(t.fh_autorizacion),
        TRUNC(t.FH_AUTORIZACION),
        CAST(t.t_ocompra AS VARCHAR2(10)),
        CAST(t.n_ocompra AS VARCHAR2(10)),
        CAST(t.aa_ocompra AS NUMBER)
    HAVING NVL(SUM(d.i_devengado), 0) != 0

UNION ALL

SELECT -- FORM 06
        t.o_ente AS Beneficiario,
        CAST(t.aa_certificado AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_certificado AS VARCHAR2(10)) AS t_formulario,
        CAST(t.n_certificado AS NUMBER) AS o_formulario,
        TRUNC(t.fh_autorizacion) AS fh_imputacion,
        TRUNC(t.FH_AUTORIZACION) AS NEW_FH_IMPUTACION,
        CAST(t.t_ocompra AS VARCHAR2(10)) AS aa_comprobante,
        CAST(t.n_ocompra AS VARCHAR2(10)) AS t_comprobante,
        CAST(t.aa_ocompra AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        SUM(d.i_devengado) AS i_devengado,
        CAST(NULL AS NUMBER(16,2)) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tanticipo_financiero' AS comentario
    FROM slu.tanticipo_financiero t
    JOIN slu.danticipo_financiero_ffi d
       ON t.aa_certificado = d.aa_ejervg
      AND t.o_certificado = d.o_certificado
    WHERE t.e_certificado = 'A'
      AND t.aa_certificado = {anio}
      AND (t.fh_autorizacion IS NULL OR t.aa_certificado != EXTRACT(YEAR FROM t.fh_autorizacion))
    GROUP BY
        t.o_ente,
        CAST(t.aa_certificado AS VARCHAR2(10)),
        CAST(t.t_certificado AS VARCHAR2(10)),
        CAST(t.n_certificado AS NUMBER),
        TRUNC(t.fh_autorizacion),
        TRUNC(t.FH_AUTORIZACION),
        CAST(t.t_ocompra AS VARCHAR2(10)),
        CAST(t.n_ocompra AS VARCHAR2(10)),
        CAST(t.aa_ocompra AS NUMBER)
    HAVING NVL(SUM(d.i_devengado), 0) != 0

UNION ALL

SELECT -- FORM 08
        t.o_beneficiario AS Beneficiario,
        CAST(t.aa_formulario AS VARCHAR2(10)) AS aa_formulario,
        CAST(t.t_formulario AS VARCHAR2(10)) AS t_formulario,
        CAST(t.o_formulario AS NUMBER) AS o_formulario,
        TRUNC(t.fh_imputacion) AS fh_imputacion,
        TRUNC(t.fh_imputacion) AS NEW_FH_IMPUTACION,
        CAST(t.aa_form_orig AS VARCHAR2(10)) AS aa_comprobante,
        CAST(t.t_form_orig AS VARCHAR2(10)) AS t_comprobante,
        CAST(t.o_form_orig AS NUMBER) AS o_comprobante,
        CAST(NULL AS VARCHAR2(50)) AS c_mediopago,
        CAST(NULL AS NUMBER(16,2)) AS i_devengado,
        SUM(d.i_pagado) AS i_pagado,
        CAST(NULL AS NUMBER(16,2)) AS ia_pago,
        'tformulario_c55' AS comentario
    FROM gs_tformulario t
    JOIN gs_dform_item d
       ON t.aa_formulario = d.aa_formulario
      AND t.t_formulario = d.t_formulario
      AND t.o_formulario = d.o_formulario
    WHERE t.t_formulario = 'C55'
      AND t.e_formulario NOT IN ('X','I','S')
      AND d.c_numcred IS NOT NULL
      AND t.aa_formulario = {anio}
      AND (t.fh_imputacion IS NULL OR t.aa_formulario != EXTRACT(YEAR FROM t.fh_imputacion))
    GROUP BY
        t.o_beneficiario,
        CAST(t.aa_formulario AS VARCHAR2(10)),
        CAST(t.t_formulario AS VARCHAR2(10)),
        CAST(t.o_formulario AS NUMBER),
        TRUNC(t.fh_imputacion),
        TRUNC(t.fh_imputacion),
        CAST(t.aa_form_orig AS VARCHAR2(10)),
        CAST(t.t_form_orig AS VARCHAR2(10)),
        CAST(t.o_form_orig AS NUMBER)
    HAVING NVL(SUM(d.i_pagado), 0) != 0
ORDER BY Beneficiario
    """
    return query_base.replace("{anio}", str(anio))

def add_subtotals(df):
    """
    Inserta sub-totales por cada Beneficiario en el DataFrame.
    Se asume que las columnas i_devengado, i_pagado y ia_pago podrían existir.
    Si alguna no existe, simplemente la ignora.
    """
    if 'BENEFICIARIO' not in df.columns:
        # Si no existe, no podemos agrupar
        return df

    group_cols = ['BENEFICIARIO']  # Agrupamos por beneficiario
    sum_cols = []
    for col in ['I_DEVENGADO', 'I_PAGADO', 'IA_PAGO']:
        if col in df.columns:
            sum_cols.append(col)

    # Si no hay columnas para resumir, retornamos sin cambios
    if not sum_cols:
        return df

    # Ordenamos por Beneficiario
    df_sorted = df.sort_values(by="BENEFICIARIO", ascending=True)

    # Dividimos en grupos y armamos un nuevo DF con subtotales
    new_rows = []
    for beneficiary, group_data in df_sorted.groupby('BENEFICIARIO', dropna=False):
        # Se añaden las filas originales
        new_rows.append(group_data)

        # Calculamos sub-totales
        subtotal_dict = {col: '' for col in group_data.columns}  # fila vacía
        subtotal_dict['BENEFICIARIO'] = 'Sub-total'
        for col in sum_cols:
            subtotal_dict[col] = group_data[col].sum()
        # Agregamos la fila de sub-total
        new_rows.append(pd.DataFrame([subtotal_dict]))

    # Concatenamos todo en un solo DataFrame
    final_df = pd.concat(new_rows, ignore_index=True)
    return final_df

def export_to_excel(dataframes, output_file):
    """
    Exporta los DataFrames (con subtotales) a un archivo Excel,
    insertando una fila celeste para cada sub-total.
    """
    # Definimos un color de relleno (celeste)
    celeste_fill = PatternFill(start_color='B0E0E6', end_color='B0E0E6', fill_type='solid')

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for year, df in dataframes.items():
            # Añadimos sub-totales
            df_con_subtotales = add_subtotals(df)

            # Exportamos a una hoja de Excel
            df_con_subtotales.to_excel(writer, sheet_name=str(year), index=False)

            # Aplicamos estilo a las filas de 'Sub-total'
            # Obtenemos la hoja
            ws = writer.sheets[str(year)]

            # Buscamos las filas donde en la columna 'BENEFICIARIO' aparezca 'Sub-total'
            # y les aplicamos el relleno celeste
            # (Columna 'BENEFICIARIO' => df_con_subtotales.columns.index('BENEFICIARIO') + 1 en Excel)
            beneficiario_col_idx = df_con_subtotales.columns.get_loc('BENEFICIARIO') + 1

            for row_idx in range(2, len(df_con_subtotales) + 2):  # Excel empieza en fila 1 para encabezado
                cell_value = ws.cell(row=row_idx, column=beneficiario_col_idx).value
                if cell_value == 'Sub-total':
                    # Aplicar estilo en toda la fila
                    for col_idx in range(1, len(df_con_subtotales.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = celeste_fill

    print(f"Archivo Excel generado correctamente: {output_file}")

def main():
    """
    Programa principal: solicita el rango de años, ejecuta las consultas y exporta los resultados a Excel.
    """
    # Validación del rango de años
    try:
        anio_inicial = int(input("Ingrese el año inicial: "))
        anio_final = int(input("Ingrese el año final: "))
        if anio_inicial > anio_final:
            print("El año inicial no puede ser mayor que el año final.")
            return
    except ValueError:
        print("Por favor, ingrese un rango de años válido.")
        return

    # Configuración de la conexión Oracle
    host = "10.17.25.52"
    sid = "preprod"
    usuario = "SLU"
    contrasena = "fsugyt8h"

    dsn_tns = cx_Oracle.makedsn(host, 1521, sid=sid)

    try:
        connection = cx_Oracle.connect(user=usuario, password=contrasena, dsn=dsn_tns)
        print("Conexión establecida correctamente.")
    except cx_Oracle.Error as e:
        print(f"Error al conectar a la base de datos: {e}")
        return

    # Nombre del archivo Excel de salida (incluye timestamp)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"resultado_query_por_anio_{timestamp}.xlsx"

    dataframes = {}  # Almacena los DataFrames por año

    try:
        cursor = connection.cursor()

        for anio in range(anio_inicial, anio_final + 1):
            sql_query = construir_query(anio)
            try:
                cursor.execute(sql_query)
                rows = cursor.fetchall()
                col_names = [desc[0] for desc in cursor.description]

                # Crear DataFrame y verificar si tiene datos
                df = pd.DataFrame(rows, columns=col_names)
                if not df.empty:
                    # Ordenar por BENEFICIARIO (si existe)
                    if 'BENEFICIARIO' in df.columns:
                        df.sort_values(by="BENEFICIARIO", ascending=True, inplace=True)

                    dataframes[anio] = df
                    print(f"Año {anio}: {len(df)} registros exportados.")
                else:
                    print(f"No se encontraron datos para el año {anio}.")

            except cx_Oracle.Error as e:
                print(f"Error al ejecutar la consulta para el año {anio}: {e}")

        cursor.close()

        # Exportar a Excel (con subtotales y filas celestes)
        if dataframes:
            export_to_excel(dataframes, output_file)
        else:
            print("No se encontraron datos para exportar a Excel.")

    except cx_Oracle.Error as e:
        print(f"Error al ejecutar las consultas: {e}")
    finally:
        connection.close()
        print("Conexión a la base de datos cerrada.")


if __name__ == "__main__":
    main()